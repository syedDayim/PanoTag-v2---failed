"""Excel export — column order per spec."""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


HEADERS = [
    "Photo Name",
    "Tag Name",
    "Pan (TL)",
    "Tilt (TL)",
    "Pan (TR)",
    "Tilt (TR)",
    "Pan (BR)",
    "Tilt (BR)",
    "Pan (BL)",
    "Tilt (BL)",
    "Confidence",
]


def export_tags_to_xlsx(
    rows: list[dict],
    out_path: str | Path,
) -> Path:
    """rows: dicts with keys matching tag export fields."""
    out_path = Path(out_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "Tag Data"
    hfill = PatternFill("solid", fgColor="1A1F35")
    hfont = Font(bold=True, color="00E5FF", name="Calibri", size=11)
    ctr = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="2A3045")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hfont
        c.fill = hfill
        c.alignment = ctr
        c.border = bdr

    cfills = {
        "TL": PatternFill("solid", fgColor="003A4A"),
        "TR": PatternFill("solid", fgColor="4A2000"),
        "BR": PatternFill("solid", fgColor="4A0000"),
        "BL": PatternFill("solid", fgColor="1A4A00"),
    }
    cfonts = {
        "TL": Font(color="00E5FF", name="Calibri"),
        "TR": Font(color="FF6B35", name="Calibri"),
        "BR": Font(color="FF4545", name="Calibri"),
        "BL": Font(color="7FFF6E", name="Calibri"),
    }

    for ri, tag in enumerate(rows, 2):
        row_data = [
            tag.get("photo", ""),
            tag.get("tag_name", ""),
            tag.get("pan_tl"),
            tag.get("tilt_tl"),
            tag.get("pan_tr"),
            tag.get("tilt_tr"),
            tag.get("pan_br"),
            tag.get("tilt_br"),
            tag.get("pan_bl"),
            tag.get("tilt_bl"),
            tag.get("conf", tag.get("confidence", 0)),
        ]
        alt = PatternFill("solid", fgColor="0D1018" if ri % 2 == 0 else "111520")
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=ri, column=col, value=val)
            c.fill = alt
            c.alignment = ctr
            c.border = bdr
            c.font = Font(name="Calibri", color="E8EAF0")
        for ci, corner in enumerate(["TL", "TR", "BR", "BL"]):
            for offset in (0, 1):
                col = 3 + ci * 2 + offset
                ws.cell(row=ri, column=col).fill = cfills[corner]
                ws.cell(row=ri, column=col).font = cfonts[corner]
        cc = ws.cell(row=ri, column=11)
        conf = float(tag.get("conf", 0) or 0)
        if conf > 0.7:
            cc.font = Font(name="Calibri", color="7FFF6E", bold=True)
        elif conf > 0.4:
            cc.font = Font(name="Calibri", color="FFB347")
        else:
            cc.font = Font(name="Calibri", color="FF4545")

    for i, w in enumerate([28, 24, 10, 10, 10, 10, 10, 10, 10, 10, 12], 1):
        from openpyxl.utils import get_column_letter

        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    wb.save(out_path)
    return out_path
