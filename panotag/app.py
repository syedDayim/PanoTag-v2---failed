"""
PanoTag — 360° Automatic Tag Extraction Tool
============================================
Loads a 360 equirectangular photo, detects physical equipment tags
using YOLOv8 + EasyOCR, converts bounding box corners to pan/tilt
angles, and exports results to Excel.

Handles very large images (24000×12000+) correctly:
  - Detection runs on a downsampled working copy (~4096px wide)
  - OCR runs on the ORIGINAL high-res crop for best accuracy
  - Coordinates are remapped back to original resolution
  - Preview uses OpenCV (no Pillow pixel limit)
"""

import sys
import threading
from pathlib import Path

# ── Raise Pillow decompression bomb limit BEFORE any Image.open() ───────────
from PIL import Image
Image.MAX_IMAGE_PIXELS = None   # we trust our own files

from PIL import ImageTk, ImageDraw, ImageFont

import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog, ttk

try:
    import cv2
    import numpy as np
except ImportError:
    print("Missing: pip install opencv-python numpy")
    sys.exit(1)

try:
    from ultralytics import YOLO
except ImportError:
    YOLO = None

try:
    import easyocr
except ImportError:
    easyocr = None

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    openpyxl = None


# ── Working resolution for detection ────────────────────────────────────────
# Detection runs on a downsampled copy; OCR always uses the original hi-res crop
DETECT_MAX_W = 4096

# Speed / precision tradeoffs (generic yolov8n.pt is COCO — not tag-specific; it
# mostly adds false boxes + cost). EasyOCR **full-image** readtext scales badly with
# width; hi-res **per-crop** OCR still uses the original image.
USE_YOLO_FOR_TAGS = False
# ~2400px proposals: much faster than 4096, still beats 1920 for small text.
TEXT_SCAN_MAX_W = 2400
# Each candidate runs a separate crop OCR — cap keeps total time bounded.
MAX_TAG_CANDIDATES = 48
MIN_REGION_CONF = 0.20
MIN_OCR_LINE_CONF = 0.20
MIN_TAG_TEXT_LEN = 1
NMS_IOU_TAGS = 0.48


# ═══════════════════════════════════════════════════════════════════════════
#  COORDINATE CONVERSION
# ═══════════════════════════════════════════════════════════════════════════

def pixel_to_pan_tilt(x, y, img_w, img_h):
    pan  = (x / img_w) * 360.0 - 180.0
    tilt = 90.0 - (y / img_h) * 180.0
    return round(pan, 4), round(tilt, 4)


def box_to_corners_pan_tilt(x1, y1, x2, y2, img_w, img_h):
    tl = pixel_to_pan_tilt(x1, y1, img_w, img_h)
    tr = pixel_to_pan_tilt(x2, y1, img_w, img_h)
    br = pixel_to_pan_tilt(x2, y2, img_w, img_h)
    bl = pixel_to_pan_tilt(x1, y2, img_w, img_h)
    return {
        "pan_tl": tl[0], "tilt_tl": tl[1],
        "pan_tr": tr[0], "tilt_tr": tr[1],
        "pan_br": br[0], "tilt_br": br[1],
        "pan_bl": bl[0], "tilt_bl": bl[1],
    }


# ═══════════════════════════════════════════════════════════════════════════
#  DETECTION ENGINE
# ═══════════════════════════════════════════════════════════════════════════

class TagDetector:

    def __init__(self, status_cb=None):
        self.status_cb = status_cb or print
        self.yolo   = None
        self.reader = None
        self._load_models()

    def _load_models(self):
        if YOLO is not None:
            self.status_cb("Loading YOLOv8 model…")
            try:
                self.yolo = YOLO("yolov8n.pt")
                self.status_cb("YOLOv8 ready ✓")
            except Exception as e:
                self.status_cb(f"YOLOv8 load failed ({e}) — OCR-only mode")

        if easyocr is not None:
            self.status_cb("Loading EasyOCR (first run ~200 MB download)…")
            try:
                self.reader = easyocr.Reader(['en'], gpu=False, verbose=False)
                self.status_cb("EasyOCR ready ✓")
            except Exception as e:
                self.status_cb(f"EasyOCR load failed ({e})")
        else:
            self.status_cb("EasyOCR not installed — install with: pip install easyocr")

    @staticmethod
    def _learned_corrections_lookup() -> dict[str, str]:
        try:
            from panotag.db import get_learned_correction_map
            return get_learned_correction_map()
        except ImportError:
            pass
        except Exception:
            return {}
        try:
            _pd = Path(__file__).resolve().parent
            if str(_pd) not in sys.path:
                sys.path.insert(0, str(_pd))
            import db as _dbmod
            return _dbmod.get_learned_correction_map()
        except Exception:
            return {}

    @staticmethod
    def _apply_learned_label(raw: str, lookup: dict[str, str]) -> str:
        s = raw.strip()
        if not s:
            return "UNKNOWN"
        if s.upper() == "UNKNOWN":
            return s
        return lookup.get(s.lower(), s)

    # ── Main pipeline ────────────────────────────────────────────────────────
    def detect(self, image_path: str, photo_stem: str | None = None) -> list[dict]:
        """
        1. Read full image via OpenCV (no pixel limit)
        2. Downsample to DETECT_MAX_W for region detection
        3. Map detected regions back to original coordinates
        4. OCR each region using the ORIGINAL hi-res crop
        5. Compute pan/tilt from original coordinates

        photo_stem: optional logical name without extension (e.g. when image_path is a temp file).
        """
        self.status_cb(f"Reading image: {Path(image_path).name}")
        img_bgr = cv2.imread(image_path)
        if img_bgr is None:
            raise ValueError(f"Cannot read image: {image_path}")

        img_rgb = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2RGB)
        orig_h, orig_w = img_rgb.shape[:2]
        self.status_cb(f"Original size: {orig_w}×{orig_h} px")

        # Downsample for detection
        if orig_w > DETECT_MAX_W:
            scale = DETECT_MAX_W / orig_w
            det_w = DETECT_MAX_W
            det_h = int(orig_h * scale)
            img_det = cv2.resize(img_rgb, (det_w, det_h), interpolation=cv2.INTER_AREA)
            self.status_cb(f"Downsampled to {det_w}×{det_h} px for detection")
        else:
            scale = 1.0
            det_w, det_h = orig_w, orig_h
            img_det = img_rgb

        # Get boxes on downsampled image
        det_boxes = self._get_boxes(img_det, det_w, det_h)
        self.status_cb(f"Found {len(det_boxes)} candidate region(s)")

        # Remap to original resolution
        orig_boxes = []
        for (x1, y1, x2, y2, conf) in det_boxes:
            ox1 = max(0, min(int(x1 / scale), orig_w - 1))
            oy1 = max(0, min(int(y1 / scale), orig_h - 1))
            ox2 = max(0, min(int(x2 / scale), orig_w - 1))
            oy2 = max(0, min(int(y2 / scale), orig_h - 1))
            orig_boxes.append((ox1, oy1, ox2, oy2, conf))

        lookup = self._learned_corrections_lookup()
        if lookup:
            self.status_cb(f"Applying {len(lookup)} learned text correction(s) from database…")

        tags = []
        photo_name = (photo_stem or Path(image_path).stem).strip() or "photo"
        for i, (x1, y1, x2, y2, conf) in enumerate(orig_boxes):
            raw_ocr, ocr_conf = self._ocr_hires_crop(img_rgb, x1, y1, x2, y2)
            if len(raw_ocr.strip()) < MIN_TAG_TEXT_LEN:
                raw_ocr = "UNKNOWN"
            resolved = self._apply_learned_label(raw_ocr, lookup)
            corners = box_to_corners_pan_tilt(x1, y1, x2, y2, orig_w, orig_h)
            blend = float(conf) * 0.5 + float(ocr_conf) * 0.5
            tags.append({
                "photo":       photo_name,
                "tag_id":      0,
                "tag_name":    resolved,
                "system_text": raw_ocr,
                "ocr_conf":    round(float(ocr_conf), 4),
                "conf":        round(blend, 3),
                "x1": x1, "y1": y1, "x2": x2, "y2": y2,
                **corners
            })
            disp = (
                f"'{resolved}'"
                if resolved == raw_ocr
                else f"'{resolved}' (OCR was '{raw_ocr}')"
            )
            self.status_cb(
                f"  [{len(tags)}] {disp}  det={conf:.2f} ocr={ocr_conf:.2f}  "
                f"pan_tl={corners['pan_tl']}  tilt_tl={corners['tilt_tl']}"
            )

        for j, t in enumerate(tags):
            t["tag_id"] = j + 1
        return tags

    # ── Region detection ─────────────────────────────────────────────────────
    def _get_boxes(self, img_rgb, w, h):
        boxes = []

        # A — YOLO (optional: default off — COCO weights are not tag/nameplate-specific)
        if USE_YOLO_FOR_TAGS and self.yolo is not None:
            self.status_cb("Running YOLOv8 detection…")
            results = self.yolo(
                img_rgb, verbose=False, conf=0.45, imgsz=min(w, 1280)
            )
            for r in results:
                for box in r.boxes:
                    x1, y1, x2, y2 = map(int, box.xyxy[0].tolist())
                    conf = float(box.conf[0])
                    bw, bh = x2 - x1, y2 - y1
                    if bw > 15 and bh > 8 and 0.1 < bw / max(bh, 1) < 25:
                        boxes.append((x1, y1, x2, y2, conf))

        # B — EasyOCR region scan on a smaller image (fast proposals; coords → full det space)
        if self.reader is not None:
            self.status_cb("Running EasyOCR region scan…")
            try:
                if w > TEXT_SCAN_MAX_W:
                    ts = TEXT_SCAN_MAX_W / w
                    sw, sh = TEXT_SCAN_MAX_W, max(1, int(h * ts))
                    scan_bgr = cv2.resize(img_rgb, (sw, sh), interpolation=cv2.INTER_AREA)
                    sx, sy = w / sw, h / sh
                else:
                    scan_bgr = img_rgb
                    sw, sh = w, h
                    sx = sy = 1.0

                ocr_raw = self.reader.readtext(
                    scan_bgr, detail=1, paragraph=False,
                    width_ths=0.9, text_threshold=0.32, low_text=0.22,
                )
                for (quad, text, conf) in ocr_raw:
                    if conf < MIN_REGION_CONF or len(text.strip()) < 1:
                        continue
                    xs = [p[0] for p in quad]
                    ys = [p[1] for p in quad]
                    x1 = int(min(xs) * sx)
                    y1 = int(min(ys) * sy)
                    x2 = int(max(xs) * sx)
                    y2 = int(max(ys) * sy)
                    pad_x = max(10, int((x2 - x1) * 0.18))
                    pad_y = max(6, int((y2 - y1) * 0.30))
                    x1 = max(0, x1 - pad_x)
                    y1 = max(0, y1 - pad_y)
                    x2 = min(w, x2 + pad_x)
                    y2 = min(h, y2 + pad_y)
                    boxes.append((x1, y1, x2, y2, conf))
            except Exception as e:
                self.status_cb(f"EasyOCR scan error: {e}")

        # C — MSER fallback
        if not boxes:
            self.status_cb("Using MSER contour fallback…")
            boxes = self._mser_detect(img_rgb, w, h)

        merged = self._nms(boxes, iou_threshold=NMS_IOU_TAGS)
        merged.sort(key=lambda b: b[4], reverse=True)
        return merged[:MAX_TAG_CANDIDATES]

    def _mser_detect(self, img_rgb, w, h):
        gray = cv2.cvtColor(img_rgb, cv2.COLOR_RGB2GRAY)
        mser = cv2.MSER_create()
        regions, _ = mser.detectRegions(gray)
        boxes = []
        for pts in regions:
            x, y, bw, bh = cv2.boundingRect(pts.reshape(-1, 1, 2))
            if bw < 20 or bh < 8 or bw > w*0.4 or bh > h*0.4:
                continue
            if 0.3 < bw / max(bh, 1) < 20:
                boxes.append((x, y, x+bw, y+bh, 0.25))
        return boxes[:MAX_TAG_CANDIDATES]

    # ── OCR on the ORIGINAL hi-res crop ──────────────────────────────────────
    def _ocr_hires_crop(self, img_rgb, x1, y1, x2, y2):
        """
        Crop from full-resolution image; return the single best OCR line and its
        confidence when clear enough — avoids merging every line in the crop.
        """
        if self.reader is None:
            return "UNKNOWN", 0.0

        crop = img_rgb[y1:y2, x1:x2]
        if crop.size == 0:
            return "UNKNOWN", 0.0

        ch, cw = crop.shape[:2]

        if cw > 800:
            factor = 800 / cw
            crop = cv2.resize(crop, (800, max(1, int(ch * factor))),
                              interpolation=cv2.INTER_AREA)
        elif cw < 80:
            factor = max(2, 160 // cw)
            crop = cv2.resize(crop, (cw * factor, ch * factor),
                              interpolation=cv2.INTER_CUBIC)

        kernel = np.array([[0, -1, 0], [-1, 5, -1], [0, -1, 0]])
        crop = cv2.filter2D(crop, -1, kernel)

        try:
            results = self.reader.readtext(
                crop, detail=1, paragraph=False,
                text_threshold=0.35, low_text=0.22, width_ths=0.9,
            )
            if not results:
                return "UNKNOWN", 0.0
            best = max(results, key=lambda r: r[2])
            text, line_conf = best[1].strip(), float(best[2])
            if len(text) < 1:
                return "UNKNOWN", line_conf
            if line_conf < MIN_OCR_LINE_CONF:
                self.status_cb(
                    f"    OCR (low conf): '{text[:40]}'  line_conf={line_conf:.2f}"
                )
            else:
                self.status_cb(f"    OCR: '{text[:40]}'  line_conf={line_conf:.2f}")
            return text, line_conf
        except Exception as e:
            self.status_cb(f"    OCR error: {e}")
            return "UNKNOWN", 0.0

    def _nms(self, boxes, iou_threshold=0.35):
        if not boxes:
            return []
        boxes = sorted(boxes, key=lambda b: b[4], reverse=True)
        kept  = []
        while boxes:
            best = boxes.pop(0)
            kept.append(best)
            boxes = [b for b in boxes if self._iou(best, b) < iou_threshold]
        return kept

    @staticmethod
    def _iou(a, b):
        ax1,ay1,ax2,ay2,_ = a
        bx1,by1,bx2,by2,_ = b
        ix1=max(ax1,bx1); iy1=max(ay1,by1)
        ix2=min(ax2,bx2); iy2=min(ay2,by2)
        inter = max(0,ix2-ix1)*max(0,iy2-iy1)
        if inter == 0: return 0
        ua = (ax2-ax1)*(ay2-ay1)+(bx2-bx1)*(by2-by1)-inter
        return inter/max(ua,1)


# ═══════════════════════════════════════════════════════════════════════════
#  ANNOTATED IMAGE RENDERER  (OpenCV-based — no Pillow pixel limit)
# ═══════════════════════════════════════════════════════════════════════════

CORNER_COLORS_BGR = {
    "TL": (255, 229, 0),    # cyan
    "TR": (53,  107, 255),  # orange
    "BR": (69,  69,  255),  # red
    "BL": (110, 255, 127),  # green
}

def render_annotated(image_path: str, tags: list[dict],
                     preview_w=1400, preview_h=700) -> Image.Image:
    img_bgr = cv2.imread(image_path)
    orig_h, orig_w = img_bgr.shape[:2]

    scale = min(preview_w / orig_w, preview_h / orig_h, 1.0)
    if scale < 1.0:
        nw = int(orig_w * scale)
        nh = int(orig_h * scale)
        img_bgr = cv2.resize(img_bgr, (nw, nh), interpolation=cv2.INTER_AREA)

    for tag in tags:
        x1 = int(tag["x1"]*scale); y1 = int(tag["y1"]*scale)
        x2 = int(tag["x2"]*scale); y2 = int(tag["y2"]*scale)
        conf  = tag["conf"]
        label = f"[{tag['tag_id']}] {tag['tag_name'][:30]}"

        if conf > 0.7:   color = (255, 229, 0)
        elif conf > 0.4: color = (0,   165, 255)
        else:            color = (80,  80,  255)

        lw = max(2, int(3*scale))
        cv2.rectangle(img_bgr, (x1,y1), (x2,y2), color, lw)

        dot_r = max(3, int(7*scale))
        for corner, (cx,cy) in [("TL",(x1,y1)),("TR",(x2,y1)),
                                  ("BR",(x2,y2)),("BL",(x1,y2))]:
            cv2.circle(img_bgr, (cx,cy), dot_r, CORNER_COLORS_BGR[corner], -1)
            cv2.circle(img_bgr, (cx,cy), dot_r, (0,0,0), 1)

        fs  = max(0.4, min(0.7, scale*2))
        lth = max(1, int(scale*2))
        tw, th = cv2.getTextSize(label, cv2.FONT_HERSHEY_SIMPLEX, fs, lth)[0]
        ly = max(th+6, y1-4)
        cv2.rectangle(img_bgr, (x1, ly-th-4), (x1+tw+4, ly+2), (0,0,0), -1)
        cv2.putText(img_bgr, label, (x1+2, ly), cv2.FONT_HERSHEY_SIMPLEX,
                    fs, color, lth, cv2.LINE_AA)

    img_rgb = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2RGB)
    return Image.fromarray(img_rgb)


# ═══════════════════════════════════════════════════════════════════════════
#  EXCEL EXPORT
# ═══════════════════════════════════════════════════════════════════════════

def export_excel(tags: list[dict], out_path: str) -> str:
    if openpyxl is None:
        out_path = out_path.replace(".xlsx", ".csv")
        with open(out_path, "w", newline="", encoding="utf-8") as f:
            f.write("photo,tag_name,pan_tl,tilt_tl,pan_tr,tilt_tr,"
                    "pan_br,tilt_br,pan_bl,tilt_bl,confidence\n")
            for t in tags:
                f.write(f"{t['photo']},{t['tag_name']},"
                        f"{t['pan_tl']},{t['tilt_tl']},"
                        f"{t['pan_tr']},{t['tilt_tr']},"
                        f"{t['pan_br']},{t['tilt_br']},"
                        f"{t['pan_bl']},{t['tilt_bl']},"
                        f"{t['conf']}\n")
        return out_path

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tag Data"

    headers = ["Photo Name","Tag Name",
               "Pan (TL)","Tilt (TL)","Pan (TR)","Tilt (TR)",
               "Pan (BR)","Tilt (BR)","Pan (BL)","Tilt (BL)","Confidence"]
    hfill = PatternFill("solid", fgColor="1A1F35")
    hfont = Font(bold=True, color="00E5FF", name="Calibri", size=11)
    ctr   = Alignment(horizontal="center", vertical="center")
    thin  = Side(style="thin", color="2A3045")
    bdr   = Border(left=thin,right=thin,top=thin,bottom=thin)

    for col,h in enumerate(headers,1):
        c=ws.cell(row=1,column=col,value=h)
        c.font=hfont; c.fill=hfill; c.alignment=ctr; c.border=bdr

    cfills = {
        "TL":PatternFill("solid",fgColor="003A4A"),
        "TR":PatternFill("solid",fgColor="4A2000"),
        "BR":PatternFill("solid",fgColor="4A0000"),
        "BL":PatternFill("solid",fgColor="1A4A00"),
    }
    cfonts = {
        "TL":Font(color="00E5FF",name="Calibri"),
        "TR":Font(color="FF6B35",name="Calibri"),
        "BR":Font(color="FF4545",name="Calibri"),
        "BL":Font(color="7FFF6E",name="Calibri"),
    }

    for ri,tag in enumerate(tags,2):
        row_data=[tag["photo"],tag["tag_name"],
                  tag["pan_tl"],tag["tilt_tl"],
                  tag["pan_tr"],tag["tilt_tr"],
                  tag["pan_br"],tag["tilt_br"],
                  tag["pan_bl"],tag["tilt_bl"],tag["conf"]]
        alt=PatternFill("solid",fgColor="0D1018" if ri%2==0 else "111520")
        for col,val in enumerate(row_data,1):
            c=ws.cell(row=ri,column=col,value=val)
            c.fill=alt; c.alignment=ctr; c.border=bdr
            c.font=Font(name="Calibri",color="E8EAF0")
        for ci,corner in enumerate(["TL","TR","BR","BL"]):
            for offset in [0,1]:
                col=3+ci*2+offset
                ws.cell(row=ri,column=col).fill=cfills[corner]
                ws.cell(row=ri,column=col).font=cfonts[corner]
        cc=ws.cell(row=ri,column=11)
        if tag["conf"]>0.7:   cc.font=Font(name="Calibri",color="7FFF6E",bold=True)
        elif tag["conf"]>0.4: cc.font=Font(name="Calibri",color="FFB347")
        else:                  cc.font=Font(name="Calibri",color="FF4545")

    for i,w in enumerate([28,24,10,10,10,10,10,10,10,10,12],1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width=w
    ws.freeze_panes="A2"
    ws.row_dimensions[1].height=22
    wb.save(out_path)
    return out_path


# ═══════════════════════════════════════════════════════════════════════════
#  GUI
# ═══════════════════════════════════════════════════════════════════════════

class PanoTagApp:
    def __init__(self, root):
        self.root       = root
        self.root.title("PanoTag — 360° Automatic Tag Extractor")
        self.root.configure(bg="#0d0f14")
        self.root.geometry("1200x750")
        self.image_path = None
        self.tags       = []
        self.detector   = None
        self.photo_img  = None
        self._build_ui()
        self._init_sqlite()
        self._init_detector()

    def _build_ui(self):
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Dark.Treeview",
            background="#151820",foreground="#e8eaf0",
            fieldbackground="#151820",rowheight=26,font=("Courier",9))
        style.configure("Dark.Treeview.Heading",
            background="#1c2030",foreground="#00e5ff",font=("Courier",9,"bold"))
        style.map("Dark.Treeview",background=[("selected","#2a3045")])

        top=tk.Frame(self.root,bg="#151820",height=52)
        top.pack(fill="x",side="top")
        top.pack_propagate(False)
        tk.Label(top,text="PANOTAG",bg="#151820",fg="#00e5ff",
                 font=("Courier",14,"bold")).pack(side="left",padx=20,pady=12)
        tk.Label(top,text="360° Automatic Tag Extractor",bg="#151820",
                 fg="#6b7590",font=("Courier",10)).pack(side="left",pady=12)

        self.export_btn=tk.Button(top,text="⬇  EXPORT EXCEL",
            bg="#7fff6e",fg="#000",relief="flat",font=("Courier",10,"bold"),
            padx=16,pady=6,state="disabled",command=self._export)
        self.export_btn.pack(side="right",padx=8,pady=10)

        self.edit_tag_btn=tk.Button(top,text="✎  EDIT TAG",
            bg="#2a3045",fg="#00e5ff",relief="flat",font=("Courier",10,"bold"),
            padx=12,pady=6,state="disabled",command=self._edit_selected_tag)
        self.edit_tag_btn.pack(side="right",padx=4,pady=10)

        self.load_btn=tk.Button(top,text="▶  LOAD PHOTO",
            bg="#00e5ff",fg="#000",relief="flat",font=("Courier",10,"bold"),
            padx=16,pady=6,command=self._load_photo)
        self.load_btn.pack(side="right",padx=4,pady=10)

        main=tk.Frame(self.root,bg="#0d0f14")
        main.pack(fill="both",expand=True)

        left=tk.Frame(main,bg="#0d0f14")
        left.pack(side="left",fill="both",expand=True,padx=(12,6),pady=12)
        tk.Label(left,text="ANNOTATED PREVIEW",bg="#0d0f14",fg="#6b7590",
                 font=("Courier",8)).pack(anchor="w")
        self.canvas=tk.Canvas(left,bg="#111520",
                              highlightthickness=1,highlightbackground="#2a3045")
        self.canvas.pack(fill="both",expand=True,pady=(4,0))
        self._draw_placeholder()

        right=tk.Frame(main,bg="#0d0f14",width=440)
        right.pack(side="right",fill="y",padx=(6,12),pady=12)
        right.pack_propagate(False)
        tk.Label(right,text="DETECTED TAGS",bg="#0d0f14",fg="#6b7590",
                 font=("Courier",8)).pack(anchor="w")

        cols=("id","tag","conf","pan_tl","tilt_tl","pan_tr","tilt_tr")
        self.tree=ttk.Treeview(right,columns=cols,show="headings",
                               style="Dark.Treeview",height=16)
        for cid,hdr,w in [("id","ID",40),("tag","Tag Name",140),
                           ("conf","Conf",50),("pan_tl","Pan TL",65),
                           ("tilt_tl","Tilt TL",65),("pan_tr","Pan TR",65),
                           ("tilt_tr","Tilt TR",65)]:
            self.tree.heading(cid,text=hdr)
            self.tree.column(cid,width=w,anchor="center")
        sb=ttk.Scrollbar(right,orient="vertical",command=self.tree.yview)
        self.tree.configure(yscrollcommand=sb.set)
        self.tree.pack(side="left",fill="both",expand=True,pady=(4,0))
        sb.pack(side="right",fill="y",pady=(4,0))
        self.tree.bind("<Double-1>", lambda e: self._edit_selected_tag())

        log_f=tk.Frame(self.root,bg="#151820",height=140)
        log_f.pack(fill="x",side="bottom")
        log_f.pack_propagate(False)
        tk.Label(log_f,text="LOG",bg="#151820",fg="#6b7590",
                 font=("Courier",8)).pack(anchor="w",padx=12,pady=(6,0))
        self.log_text=tk.Text(log_f,bg="#151820",fg="#6b7590",
                              font=("Courier",9),relief="flat",
                              state="disabled",wrap="word")
        self.log_text.pack(fill="both",expand=True,padx=12,pady=(2,8))
        for tag,col in [("info","#00e5ff"),("ok","#7fff6e"),
                         ("warn","#ffb347"),("error","#ff4545")]:
            self.log_text.tag_config(tag,foreground=col)

        self.status_var=tk.StringVar(value="Ready — load a photo to begin")
        tk.Label(self.root,textvariable=self.status_var,bg="#0d0f14",
                 fg="#6b7590",font=("Courier",9),anchor="w").pack(
                     fill="x",side="bottom",padx=12,pady=4)

    def _draw_placeholder(self):
        self.canvas.update_idletasks()
        w=self.canvas.winfo_width() or 600
        h=self.canvas.winfo_height() or 400
        self.canvas.delete("all")
        self.canvas.create_text(w//2,h//2,
            text="Load a 360° panoramic photo to begin",
            fill="#2a3045",font=("Courier",12))

    def _init_sqlite(self):
        try:
            from panotag.db import init_db
            init_db()
        except ImportError:
            try:
                _pd = Path(__file__).resolve().parent
                if str(_pd) not in sys.path:
                    sys.path.insert(0, str(_pd))
                import db as _dbmod
                _dbmod.init_db()
            except Exception as e:
                self._log(f"Database init skipped: {e}", "warn")
        except Exception as e:
            self._log(f"Database init: {e}", "warn")

    def _init_detector(self):
        self._log("Initialising models…","info")
        def _init():
            self.detector=TagDetector(status_cb=self._log_safe)
            self._log_safe("Models ready — load a photo ✓","ok")
        threading.Thread(target=_init,daemon=True).start()

    def _load_photo(self):
        path=filedialog.askopenfilename(
            title="Select 360° panoramic photo",
            filetypes=[("Images","*.jpg *.jpeg *.png *.tif *.tiff"),
                       ("All files","*.*")])
        if not path: return
        self._load_photo_path(path)

    def _load_photo_path(self, path):
        self.image_path=path
        self.tags=[]
        self._clear_table()
        self.edit_tag_btn.config(state="disabled")
        self._show_status(f"Loaded: {Path(path).name}")
        self._log(f"Loaded: {path}","info")
        self._show_raw_preview(path)
        self._run_detection()

    def _show_raw_preview(self, path):
        try:
            img_bgr=cv2.imread(path)
            img_rgb=cv2.cvtColor(img_bgr,cv2.COLOR_BGR2RGB)
            self.canvas.update_idletasks()
            cw=self.canvas.winfo_width()  or 700
            ch=self.canvas.winfo_height() or 400
            oh,ow=img_rgb.shape[:2]
            scale=min(cw/ow,ch/oh)
            nw,nh=int(ow*scale),int(oh*scale)
            small=cv2.resize(img_rgb,(nw,nh),interpolation=cv2.INTER_AREA)
            pil=Image.fromarray(small)
            self.photo_img=ImageTk.PhotoImage(pil)
            self.canvas.delete("all")
            self.canvas.create_image(cw//2,ch//2,anchor="center",
                                     image=self.photo_img)
        except Exception as e:
            self._log(f"Preview error: {e}","warn")

    def _run_detection(self):
        if self.detector is None:
            self._log("Models still loading, please wait…","warn")
            self.root.after(1500,self._run_detection)
            return
        self._show_status("Detecting tags — please wait…")
        self.load_btn.config(state="disabled")
        self.export_btn.config(state="disabled")
        self.edit_tag_btn.config(state="disabled")
        stem = Path(self.image_path).stem
        def _detect():
            try:
                tags=self.detector.detect(self.image_path, stem)
                self.root.after(0,lambda:self._on_done(tags))
            except Exception as e:
                self.root.after(0,lambda:self._on_error(str(e)))
        threading.Thread(target=_detect,daemon=True).start()

    def _on_done(self,tags):
        self.tags=tags
        self._populate_table(tags)
        self._show_annotated()
        self._show_status(f"Done — {len(tags)} tag(s) detected")
        self._log(f"Detection complete: {len(tags)} tag(s)","ok")
        self.load_btn.config(state="normal")
        if tags:
            self.export_btn.config(state="normal")
            self.edit_tag_btn.config(state="normal")

    def _on_error(self,err):
        self._log(f"Error: {err}","error")
        self._show_status(f"Error: {err}")
        self.load_btn.config(state="normal")

    def _show_annotated(self):
        if not self.image_path or not self.tags: return
        try:
            self.canvas.update_idletasks()
            cw=self.canvas.winfo_width()  or 700
            ch=self.canvas.winfo_height() or 400
            ann=render_annotated(self.image_path,self.tags,
                                 preview_w=cw*2,preview_h=ch*2)
            ann.thumbnail((cw,ch),Image.LANCZOS)
            self.photo_img=ImageTk.PhotoImage(ann)
            self.canvas.delete("all")
            self.canvas.create_image(cw//2,ch//2,anchor="center",
                                     image=self.photo_img)
        except Exception as e:
            self._log(f"Annotation error: {e}","warn")

    def _populate_table(self,tags):
        self._clear_table()
        for tag in tags:
            self.tree.insert("","end",values=(
                tag["tag_id"],tag["tag_name"][:28],
                f"{tag['conf']:.2f}",
                tag["pan_tl"],tag["tilt_tl"],
                tag["pan_tr"],tag["tilt_tr"],
            ))

    def _edit_selected_tag(self):
        if not self.tags:
            return
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo(
                "Edit tag", "Select a row in the tag table first (or double-click a row).",
            )
            return
        vals = self.tree.item(sel[0], "values")
        if not vals:
            return
        try:
            tid = int(vals[0])
        except (TypeError, ValueError):
            return
        tag = next((t for t in self.tags if t["tag_id"] == tid), None)
        if tag is None:
            return
        raw = tag.get("system_text", tag["tag_name"])
        cur = tag["tag_name"]
        new = simpledialog.askstring(
            "Edit tag text",
            f"OCR reading was: {raw}\n\nCorrect label:",
            initialvalue=cur,
            parent=self.root,
        )
        if new is None:
            return
        new = new.strip()
        if not new:
            messagebox.showwarning("Edit tag", "Label cannot be empty.")
            return
        tag["tag_name"] = new
        try:
            from panotag.db import remember_ocr_correction
            remember_ocr_correction(raw, new)
        except ImportError:
            try:
                _pd = Path(__file__).resolve().parent
                if str(_pd) not in sys.path:
                    sys.path.insert(0, str(_pd))
                import db as _dbmod
                _dbmod.remember_ocr_correction(raw, new)
            except Exception as e:
                self._log(f"Could not save correction to DB: {e}", "warn")
        except Exception as e:
            self._log(f"Could not save correction to DB: {e}", "warn")
        self._populate_table(self.tags)
        self._show_annotated()
        self._log(f"Tag #{tid} set to '{new}' (saved for next run if OCR repeats '{raw}')", "ok")

    def _clear_table(self):
        for row in self.tree.get_children(): self.tree.delete(row)

    def _export(self):
        if not self.tags:
            messagebox.showwarning("No data","No tags to export yet.")
            return
        default=Path(self.image_path).stem+"_tags.xlsx"
        out=filedialog.asksaveasfilename(
            defaultextension=".xlsx",initialfile=default,
            filetypes=[("Excel","*.xlsx"),("CSV","*.csv"),("All","*.*")])
        if not out: return
        try:
            saved=export_excel(self.tags,out)
            self._log(f"Exported → {saved}","ok")
            messagebox.showinfo("Exported",f"Saved:\n{saved}")
        except Exception as e:
            self._log(f"Export error: {e}","error")
            messagebox.showerror("Export failed",str(e))

    def _show_status(self,msg): self.status_var.set(msg)

    def _log(self,msg,level="info"):
        self.log_text.config(state="normal")
        self.log_text.insert("end",f"› {msg}\n",level)
        self.log_text.see("end")
        self.log_text.config(state="disabled")

    def _log_safe(self,msg,level="info"):
        self.root.after(0,lambda:self._log(msg,level))


# ═══════════════════════════════════════════════════════════════════════════
def main():
    root=tk.Tk()
    app=PanoTagApp(root)
    if len(sys.argv)>1 and Path(sys.argv[1]).exists():
        root.after(2500,lambda:app._load_photo_path(sys.argv[1]))
    root.mainloop()

if __name__=="__main__":
    main()
